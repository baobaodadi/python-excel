from openpyxl import Workbook,load_workbook
from datetime import datetime
wb=load_workbook("kaoqin.xlsx")
ws = wb.active
first_column = ws['A']
mydict=dict()
name=["陈国威","段静远","高媛","耿博","胡冰薇",
      "刘曼玲","卢文青","毛晓英","钱倩","孙楠楠",
      "武忠震","王琦","杨融","张羽薇","朱江",
      "邹鑫","胡文博","毛晓英"]
for var in ws.iter_rows():
    if var[0].value in name:
        mydict[var[0].value] = []

for var in ws.iter_rows():
    if var[0].value in name:
        mydict[var[0].value].append(str(var[1].value))

duration=['2018-10-15','2018-10-16','2018-10-17','2018-10-18','2018-10-19','2018-10-20','2018-10-21','2018-10-22','2018-10-23','2018-10-24']
final=dict()

for var1 in mydict.keys():
    final[var1] = dict()
    for var2 in mydict[var1]:
        for var3 in duration:
            ATime, BTime=var2.split(' ')
            if var3==ATime:
                if final[var1].get(var3):
                    final[var1][var3].append(var2)
                else:
                    final[var1][var3]=[]
                    final[var1][var3].append(var2)

sheet=wb.create_sheet('统计结果')
wb.save(filename='kaoqin.xlsx')
temp = duration.copy()
temp.insert(0, '')
sheet.append(temp)
for var1 in final.keys():
    tempA=[var1,'','','','','','','','','','']
    for var2 in final[var1]:
          tempA.insert(duration.index(var2)+1,'上班时间：'+final[var1][var2][0]+'下班时间：'+final[var1][var2][-1])
    sheet.append(tempA)

wb.save('kaoqin.xlsx')
print("保存成功")
